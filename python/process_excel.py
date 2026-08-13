import openpyxl
import re
from datetime import datetime
import os

txt_file = r"C:\Users\w1889\Desktop\处理信息.txt"
excel_file = r"C:\Users\w1889\Desktop\test.xlsx"

headers = ['序号', '时间表', '时间', '地点', '原因', '处理部门', '处理人', '来源', '维修部门', '维修人', '维修内容', '备注']

def parse_text(text):
    text = text.strip()
    if not text:
        return None

    # Match source e.g. "企业报单:..." -> source "企业"
    source = ""
    if ":" in text or "：" in text:
        separator = ":" if ":" in text else "："
        source = text.split(separator)[0].strip()
        if source.endswith("报单"):
            source = source[:-2]
        content_part = text.split(separator, 1)[1].strip()
    else:
        content_part = text

    # Split by @ for department and person
    dept_person = ""
    if "@" in content_part:
        parts = content_part.split("@")
        main_content = parts[0].strip()
        dept_person = parts[1].strip()
    else:
        main_content = content_part

    # Parse dept and person
    repair_dept = ""
    repair_person = ""
    if "-" in dept_person:
        dp_parts = dept_person.split("-")
        repair_dept = dp_parts[0].strip()
        repair_person = dp_parts[1].strip()
    else:
        repair_dept = dept_person

    # Parse location and reason
    loc_match = re.match(r"(.*?(?:楼|单元|室|\d{3,4}))(.*)", main_content)
    location = ""
    reason = ""
    if loc_match:
        location = loc_match.group(1).strip()
        reason = loc_match.group(2).strip()
        if reason.startswith("需要"):
            reason = reason[2:]
    else:
        reason = main_content

    return {
        "来源": source,
        "地点": location,
        "原因": reason,
        "维修部门": repair_dept,
        "维修人": repair_person
    }

def process():
    if not os.path.exists(txt_file):
        print(f"File not found: {txt_file}")
        return

    with open(txt_file, 'r', encoding='utf-8') as f:
        content = f.read()

    if not content.strip():
        print("No content to parse")
        return

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except:
        wb = openpyxl.Workbook()
        ws = wb.active

    # check headers
    is_empty = (ws.max_row == 1 and not ws.cell(row=1, column=1).value)
    if is_empty:
        ws.append(headers)

    # compute new index
    new_index = 1
    if not is_empty:
        last_row = ws.max_row
        last_index_val = ws.cell(row=last_row, column=1).value
        try:
            new_index = int(last_index_val) + 1
        except:
            new_index = last_row

    now = datetime.now()
    date_str = now.strftime("%Y.%m.%d")
    time_str = now.strftime("%H:%M")

    # Process each line
    added_count = 0
    for line in content.splitlines():
        line = line.strip()
        if not line:
            continue

        parsed = parse_text(line)
        if not parsed:
            continue

        date_id = now.strftime("%Y%m%d") + str(new_index)

        row_data = [
            new_index,
            date_id,
            f"{date_str} {time_str}",
            parsed["地点"],
            parsed["原因"],
            "",
            "",
            parsed["来源"],
            parsed["维修部门"],
            parsed["维修人"],
            parsed["原因"],
            ""
        ]

        ws.append(row_data)
        new_index += 1
        added_count += 1

    try:
        wb.save(excel_file)
        print(f"Success: {added_count} records appended to test.xlsx.")
    except PermissionError:
        print("Error: Please close the test.xlsx file in Excel before running the script.")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

if __name__ == '__main__':
    process()
